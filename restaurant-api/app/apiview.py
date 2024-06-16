import random
import io
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.lib.pagesizes import A4
import ast
from barcode.writer import ImageWriter
import barcode
import base64
from calendar import month_abbr, month_name
from datetime import timedelta
import xlwt
from django.http import HttpResponse
from openpyxl import load_workbook
import os
from PIL import Image
from io import BytesIO
from django.core.files import File
import operator
import functools
import collections
from collections import OrderedDict
from datetime import datetime
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from rest_framework.generics import CreateAPIView
from rest_framework import status
from rest_framework.authtoken.models import Token
from django.contrib.auth import authenticate, login
from django.core.exceptions import ObjectDoesNotExist
from django.contrib.auth import get_user_model

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from django.core.mail import EmailMessage
from . import models, serializers
from django.core.exceptions import ValidationError
from django.shortcuts import get_object_or_404
from django.db.models import Q

import json


def convert_to_number(input_value):
    if not input_value == None:
        try:
            return float(input_value)
        except ValueError:
            return 0
    else:
        return 0
    
def CHECK_IN_PLAN_AND_RESPONSE(user, data, **args):
    if user.is_plan:
        return Response('End Plan or No Purchase Plan')
    else:
        return Response(data=data, **args)

    print('User is in Plan')


class LoginView(APIView):

    permission_classes = [AllowAny]

    def post(self, request, *args, **kwargs):

        print(request.data)

        username_or_email = request.data['username']
        password = request.data['password']
        device_unique = request.data.get("unique_id", None)
        device_name = request.data.get("device_name", None)
        acc_type = request.data.get("acc_type", "Admin")

        # user devices
        device = request.data.get('device', None)

        user = None
        if '@' in username_or_email:
            b = models.User.objects.get(email=username_or_email)
            user = authenticate(username=b.username, password=password)

        else:
            user = authenticate(username=username_or_email, password=password)

        check_acc_type = models.User.objects.get(
            username=user.username, acc_type=acc_type)

        if not user or not check_acc_type:
            return Response({'error': 'Invalid Credentials'},
                            status=status.HTTP_401_UNAUTHORIZED)

        # user = serializer.validated_data['user']
        token, created = Token.objects.get_or_create(user=user)

        device_count = models.Device.objects.filter(user=user).count()

        # if device_count >= user.device_limit:
        #     raise ValidationError('Device Limit Exceeded')
        # else:
        #     models.Device.objects.create(
        #         user=user, unique_id=device_unique, device_name=device_name, acc_type=acc_type)

        # add custom data to response
        response_data = {
            'token': token.key,
            'user_id': user.id,
            'username': user.username,
            # add additional data fields here as needed
        }

        return Response(response_data)


class CreateUserApiView(CreateAPIView):

    permission_classes = [AllowAny]
    serializer_class = serializers.CreateUserSerializer

    def post(self, request):
        print(request.data)

        # device_unique = request.data.get("unique_id", None)
        # device_name = request.data.get("device_name", None)
        acc_type = request.data.get("acc_type", "Admin")

        serializers = self.get_serializer(data=request.data)
        serializers.is_valid(raise_exception=True)
        self.perform_create(serializers)

        headers = self.get_success_headers(serializers.data)
        user = models.User.objects.get(username=request.data.get('username'))
        # models.Device.objects.create(
        #     user=user, unique_id=device_unique, device_name=device_name, acc_type=acc_type)

        token = Token.objects.create(user=serializers.instance)
        token_data = {'token': token.key}

        return Response(
            {**serializers.data, **token_data},
            status=status.HTTP_201_CREATED,
            headers=headers)


class CreateCompany(APIView):

    def post(self, request):
        name = request.data.get('name', None)
        email = request.data.get('email', None)
        phoneno = request.data.get('phoneno', None)
        address = request.data.get('address', None)
        logo = request.data.get('logo', None)

        user = get_user_model().objects.get(username=request.user)
        if user.acc_type == 'Admin':
            models.CompanyProfile.objects.create(
                name=name, email=email, phoneno=phoneno, address=address, logo=logo)

        return Response(status=status.HTTP_201_CREATED)

    def get(self, request):
        data = models.CompanyProfile.objects.last()
        s = serializers.CompanyProfileSerializer(data)

        return Response(s.data)

    def put(self, request):
        data = models.CompanyProfile.objects.last()

        name = request.data.get('name', data.name)
        email = request.data.get('email', data.email)
        phoneno = request.data.get('phoneno', data.phoneno)
        address = request.data.get('address', data.address)
        logo = request.data.get('logo', None)

        print(logo, "this line logo is what that")

        if not logo == None:
            data.logo = logo

        data.name = name
        data.email = email
        data.phoneno = phoneno
        data.address = address

        data.save()

        s = serializers.CompanyProfileSerializer(data)

        return Response(s.data)


class AccountsAPIView(APIView):

    def get(self, request):
        data = models.User.objects.all()
        s = serializers.ProfileSerializer(data, many=True)

        return Response(s.data)

    def put(self, request):
        username = request.data.get('username', None)
        user = models.User.objects.get(username=username)
        user.phoneno = request.data.get('phoneno', user.phoneno)
        user.acc_type = request.data.get('acc_type', user.acc_type)

        password = request.data.get('password', None)

        if not password == None:
            user.set_password(password)

        user.save()

        return Response(status=status.HTTP_201_CREATED)

    def delete(self, request):
        username = request.GET.get('username', None)
        user = models.User.objects.get(username=username)

        if request.user.acc_type == "Admin":
            user.delete()
            return Response(status=status.HTTP_200_OK)


class KitchenAPIView(APIView):
    # permission_classes = [AllowAny]

    def get(self, request):
        data = models.Kitchen.objects.all()

        s = serializers.KitchenSerializer(data, many=True)

        return Response(s.data)

    def post(self, request):

        name = request.data.get("name", None)
        description = request.data.get("description", None)

        models.Kitchen.objects.create(name=name, description=description)

        return Response(status=status.HTTP_201_CREATED)

    def delete(self, request):
        id = request.GET.get('id', None)
        kt = models.Kitchen.objects.get(id=id)
        kt.delete()

        return Response(status=status.HTTP_200_OK)

    def put(self, request):
        id = request.data.get('id', None)

        kt = models.Kitchen.objects.get(id=id)
        kt.name = request.data.get('name', kt.name)
        kt.description = request.data.get('description', kt.description)

        kt.save()

        return Response(status=status.HTTP_200_OK)


class password_recovery(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        username = request.data.get('username', None)
        user = get_user_model().objects.get(username=username)
        otp = random.randint(100000, 999999)
        models.OTP.objects.create(user=user, otp=otp)

        subject = "Perfect Solution Password Recovery"

        html_content = '''<!DOCTYPE html>
<html lang="en">
  <head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>Reset Password</title>
  </head>
  <body>
    <p>Hi ''' + username + '''</p>
    <h1>Perfect Solutions</h1>
    <p>
      We received a request to the reset password associated with your Perfect
      Solutions' POS account.
    </p>
    <p>Here is your OTP code:<br /></p>
    <h1 style="width:100%; text-align:center">'''+str(otp)+'''</h1>
    <p>Thank you.</p>
  </body>
</html>
        '''
        SENDEMAIL = EmailMessage(subject, html_content, 'perfectsolutionpos@gmail.com', [
                                 user.email], headers={'Message-ID': user.id})
        SENDEMAIL.content_subtype = "html"
        SENDEMAIL.send()
        # TODO: Send the password reset link to the user's email address
        # email = ...
        # send_email(email, uid, token)

        return Response({"email": user.email}, status=status.HTTP_200_OK)


class CheckOTPandChangePassword(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        username = request.GET.get('username', None)
        otp = request.GET.get('otp', None)
        print(otp)
        user = get_user_model().objects.get(username=username)
        otp = models.OTP.objects.get(user=user, otp=otp)
        if otp is not None:
            return Response(status=status.HTTP_200_OK)
        else:
            return Response(status=status.HTTP_400_BAD_REQUEST)

    def post(self, request):
        username = request.data.get('username', None)
        otp = request.data.get('otp', None)
        password = request.data.get('password', None)
        oldpassword = request.data.get('oldpassword', None)

        if oldpassword != None:
            user = get_user_model().objects.get(username=request.user)

            if user.check_password(oldpassword):
                user.set_password(password)
                user.save()
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_400_BAD_REQUEST)

        else:
            user = get_user_model().objects.get(username=username)
            m_otp = models.OTP.objects.get(user=user, otp=otp)

            user.set_password(password)
            user.save()
            m_otp.delete()

            return Response(status=status.HTTP_200_OK)


class ProfileUpdate(APIView):

    def put(self, request):

        print(request.data)
        user = get_user_model().objects.get(username=request.user)
        name = request.data.get('name', user.name)
        email = request.data.get('email', user.email)
        phone = request.data.get('phone', user.phoneno)
        address = request.data.get('address', user.address)

        if name is not None:
            user.name = name
        if email is not None:
            user.email = email
        if phone is not None:
            user.phoneno = phone

        if address is not None:
            user.address = address

        user.save()
        s = serializers.ProfileSerializer(user)

        return Response(status=status.HTTP_201_CREATED, data=s.data)


class Category(APIView):
    # permission_classes = [AllowAny]

    def get(self, request):
        data = models.Category.objects.all()
        s = serializers.CategorySerializer(data, many=True)

        return Response(s.data)

    def post(self, request):
        models.Category.objects.create(title=request.data['title'])

        return Response(status=status.HTTP_201_CREATED)

    def delete(self, request):
        id = request.GET.get('id', None)
        user = get_user_model().objects.get(username=request.user)
        ct = models.Category.objects.get(id=id)
        ct.delete()

        return Response(status=status.HTTP_200_OK)

    def put(self, request):
        id = request.data.get('id', None)

        ct = models.Category.objects.get(id=id)
        ct.title = request.data.get('title', ct.title)
        ct.show = request.data.get('show', ct.show)

        print(request.data.get('show'), ct.show)

        ct.save()

        return Response(status=status.HTTP_200_OK)


def compress_image(image):
    im = Image.open(image)
    size = File(image).size
    if size > 0.2*512*512:
        print('Compressing in Progress')
        if im.mode != 'RGB':
            im = im.convert('RGB')
        im_io = BytesIO()
        im.save(im_io, 'jpeg', quality=8, optimize=True)
        print('Compressing Completed')
        new_image = File(im_io, name=image.name)
        return new_image
    return image


class Product(APIView):
    # permission_classes = [AllowAny]

    def get(self, request):

        expiry_filter_type = request.GET.get('expiry_filter_type', None)
        expiry_filter_day = request.GET.get(
            'expiry_filter_day', None)  # eg. 10 days, 5 days

        if not expiry_filter_day == None:
            data = models.Product.objects.filter(expiry_date__range=[datetime.now(
            ), datetime.now() + timedelta(days=int(expiry_filter_day))])
            s = serializers.ProductSerializer(data, many=True)

            return Response(s.data)

        if expiry_filter_type == 'expired':
            data = models.Product.objects.filter(
                expiry_date__lt=datetime.now())
        elif expiry_filter_type == 'not_expired':
            data = models.Product.objects.filter(
                expiry_date__gt=datetime.now())
        # filter this week expiry date
        elif expiry_filter_type == 'this_week':
            data = models.Product.objects.filter(
                expiry_date__range=[datetime.now(), datetime.now() + timedelta(days=7)])
        else:
            data = models.Product.objects.all()

        s = serializers.ProductSerializer(data, many=True)
        return Response(s.data)

    def post(self, request):
        name = request.data.get('name', None)
        price = request.data.get('price', None)
        cost = request.data.get('cost', 0)
        qty = request.data.get('qty', None)
        barcode = request.data.get('barcode', None)
        description = request.data.get('description', None)
        category = models.Category.objects.get(id=request.data['category'])
        pic = request.data.get('pic', 'null')
        supplier_name = request.data.get('supplier_name', None)

        expiry_date = request.data.get('expiry_date', None)

        extraprice = request.data.get('extraprice', None)  # '4000, 5000, 6000'
        unit = request.data.get('unit', 0)
        kitchen = request.data.get('kitchen', None)
        totalunit = int(unit) * int(qty)

        isunit = False
        if int(unit) >= 1:
            isunit = True

        kt = models.Kitchen.objects.get(id=kitchen)

        md = models.Product.objects.create(
            name=name,  pic=pic, price=price,
            cost=cost, qty=qty, description=description,
            barcode=barcode,
            category=category, expiry_date=expiry_date,
            unit=unit, totalunit=totalunit, isUnit=isunit,
            kitchen=kt,
        )

        if not extraprice == None:
            extraprice = extraprice.split(',')
            for p in extraprice:
                models.ProductPrice.objects.create(pdid=md, extraprice=p)

        if not supplier_name == None:
            try:
                supp = models.Supplier.objects.get(name=supplier_name)
                supp.products.add(md)
                supp.save()
            except ObjectDoesNotExist:
                supp = models.Supplier.objects.create(name=supplier_name)
                supp.products.add(md)
                supp.save()

        if not pic == 'null':
            md.pic = compress_image(pic)
            md.save()

        return Response(status=status.HTTP_201_CREATED)

    def put(self, request):
        id = request.data['id']

        pic = request.data.get('pic', 'null')

        print(request.data)

        description = request.data.get('description', None)

        PRODUCTS = models.Product.objects.get(id=id)
        PRODUCTS.name = request.data.get('name', PRODUCTS.name)
        PRODUCTS.price = request.data.get('price', PRODUCTS.price)
        PRODUCTS.qty = request.data.get('qty', PRODUCTS.qty)
        PRODUCTS.cost = request.data.get('cost', PRODUCTS.cost)
        # PRODUCTS.date = date
        PRODUCTS.description = request.data.get(
            'description', PRODUCTS.description)

        category = models.Category.objects.get(
            id=request.data.get('category', PRODUCTS.category.id))

        PRODUCTS.category = category
        PRODUCTS.barcode = request.data.get('barcode', PRODUCTS.barcode)
        PRODUCTS.expiry_date = request.data.get(
            'expiry_date', PRODUCTS.expiry_date)
        PRODUCTS.unit = request.data.get('unit', PRODUCTS.unit)
        PRODUCTS.totalunit = int(PRODUCTS.unit) * int(PRODUCTS.qty)

        kt = models.Kitchen.objects.get(
            id=request.data.get('kitchen', PRODUCTS.kitchen.id)
        )

        PRODUCTS.kitchen = kt
        extraprice = request.data.get('extraprice', None)  # '4000, 5000, 6000'

        if not extraprice == None:
            # delete all releated product from ProductPrice
            models.ProductPrice.objects.filter(pdid=PRODUCTS).delete()

            extraprice = extraprice.split(',')
            for p in extraprice:
                models.ProductPrice.objects.create(pdid=PRODUCTS, extraprice=p)

        try:
            img = Image.open(pic)
        except FileNotFoundError:
            pic = 'null'

        if not pic == 'null':
            PRODUCTS.pic = compress_image(pic)

        PRODUCTS.save()

        return Response(status=status.HTTP_201_CREATED)

    def delete(self, request):
        print(request.data)
        id = request.data['id']

        PRODUCTS = models.Product.objects.get(id=id)
        PRODUCTS.delete()
        return Response(status=status.HTTP_201_CREATED)


def calculateFoodQtyByProductUseUnit():
    food = models.Food.objects.all()

    for f in food:
        integrient = f.integrient.all()

        list = []

        for i in integrient:
            pd = models.Product.objects.get(id=i.product.id)
            print(pd.name, pd.qty, pd.totalunit, i.useunit)

            list.append(int(int(pd.totalunit) / int(i.useunit)
                            ))
        print("One Food .....................................")
        print(list)
        # get mininum value from list
        if list == []:
            pass
        else:
            f.qty = min(list)
            f.save()

            print(f.qty)


class FoodAPIView(APIView):
    def get(self, request):
        calculateFoodQtyByProductUseUnit()
        food = models.Food.objects.all()
        s = serializers.FoodSerializer(food, many=True)

        return Response(s.data)

    def post(self, request):
        name = request.data.get("name", "Unkown")
        price = request.data.get("price", 0)
        qty = request.data.get("qty", 0)
        description = request.data.get('description', None)
        category = request.data.get('category', 0)

        pic = request.data.get('pic', 'null')
        kitchen = request.data.get('kitchen', None)
        integrient = request.data.get('integrient', [])  # pdid , unit
        avaliable = request.data.get('avaliable', True)

        inte_raw = json.loads(integrient)

        kt = models.Kitchen.objects.get(id=kitchen)

        if avaliable == 'true':
            avaliable = True
        else:
            avaliable = False

        categ = models.Category.objects.get(id=category)

        # Customer.sales.add(S)
        FOOD = models.Food.objects.create(name=name, price=price,
                                          qty=qty, description=description,
                                          isavaliable=avaliable,
                                          category=categ, kitchen=kt)

        if not pic == 'null':
            FOOD.pic = compress_image(pic)
            FOOD.save()

        foodcost = 0

        for product in inte_raw:
            pd = models.Product.objects.get(id=product['id'])
            fdinte = models.FoodIntegrient.objects.create(
                product=pd, useunit=product['useunit'])

            foodcost += int(float(fdinte.useunit)) / \
                int(float(pd.unit)) * int(float(pd.cost))

            FOOD.integrient.add(fdinte)
            FOOD.save()

        totalProfit = int(FOOD.price) - int(foodcost)

        FOOD.profit = totalProfit

        FOOD.save()

        return Response(status=status.HTTP_201_CREATED)

    def put(self, request):

        id = request.data.get('id', None)

        integrient = request.data.get('integrient', None)

        print(request.data.get('id'))

        FOOD = models.Food.objects.get(id=id)

        FOOD.name = request.data.get("name", FOOD.name)
        FOOD.price = request.data.get("price", FOOD.price)
        FOOD.qty = request.data.get("qty", FOOD.qty)
        FOOD.description = request.data.get('description', FOOD.description)
        category = request.data.get("category", FOOD.category.id)
        cat = models.Category.objects.get(id=category)
        FOOD.category = cat

        pic = request.data.get('pic', 'null')
        kitchen = request.data.get('kitchen', FOOD.kitchen.id)
        kt = models.Kitchen.objects.get(id=kitchen)
        FOOD.kitchen = kt

        if not pic == 'null':
            FOOD.pic = compress_image(pic)

        if request.data.get('avaliable', None) == 'true':
            FOOD.isavaliable = True
        else:
            FOOD.isavaliable = False

        FOOD.save()

        if not integrient == None:
            inte_raw = json.loads(integrient)

            FOOD.integrient.clear()
            foodcost = 0
            for product in inte_raw:
                pd = models.Product.objects.get(id=product['id'])
                fdinte = models.FoodIntegrient.objects.create(
                    product=pd, useunit=product['useunit'])
                foodcost += int(float(fdinte.useunit)) / \
                    int(float(pd.unit)) * int(float(pd.cost))
                FOOD.integrient.add(fdinte)
                FOOD.save()

            totalProfit = int(FOOD.price) - int(foodcost)

            FOOD.profit = totalProfit
            FOOD.save()

        return Response(status=status.HTTP_201_CREATED)

    def delete(self, request):
        ids = request.GET.get('ids', None)
        allid = json.loads(ids)
        for id in allid:
            FOOD = models.Food.objects.get(id=id)
            FOOD.delete()
        return Response(status=status.HTTP_201_CREATED)


class FloorAPIView(APIView):
    def get(self, request):
        data = models.Floor.objects.all()
        s = serializers.FloorSerializer(data, many=True)

        return Response(s.data)

    def post(self, request):
        name = request.data.get('name', None)
        models.Floor.objects.create(name=name)

        return Response(status=status.HTTP_201_CREATED)

    def put(self, request):
        id = request.data.get('id', None)
        name = request.data.get('name', None)

        floor = models.Floor.objects.get(id=id)
        floor.name = name
        floor.save()

        return Response(status=status.HTTP_200_OK)

    def delete(self, request):
        id = request.GET.get('id', None)
        floor = models.Floor.objects.get(id=id)
        floor.delete()

        return Response(status=status.HTTP_200_OK)


class TableAPIView(APIView):
    def get(self, request):
        data = models.Table.objects.all()
        s = serializers.TableSerializer(data, many=True)

        return Response(s.data)

    def post(self, request):
        name = request.data.get('name', None)
        floor = models.Floor.objects.get(id=request.data.get('floor', None))
        models.Table.objects.create(name=name, floor=floor)

        return Response(status=status.HTTP_201_CREATED)

    def put(self, request):
        id = request.data.get('id', None)
        name = request.data.get('name', None)
        floor = models.Floor.objects.get(id=request.data.get('floor', None))

        table = models.Table.objects.get(id=id)
        table.name = name
        table.floor = floor
        table.save()

        return Response(status=status.HTTP_200_OK)

    def delete(self, request):
        id = request.GET.get('id', None)
        table = models.Table.objects.get(id=id)
        table.delete()

        return Response(status=status.HTTP_200_OK)


def handle_foodOrder_element(Order, item, qty):
    total_price = int(qty) * int(item.price)

    try:
        orderelement = Order.food_orders.get(
            food=item, isCooking=False, isComplete=False)
        orderelement.qty = int(orderelement.qty) + int(qty)
        orderelement.total_price = int(item.price) * int(orderelement.qty)
        orderelement.isComplete = False

        # If the quantity becomes zero, delete the order element
        if orderelement.qty == 0:
            orderelement.delete()
        else:
            orderelement.save()
    except ObjectDoesNotExist:
        # If the quantity is zero, don't create a new order element
        if qty != 0:
            orderelement = models.FoodOrder.objects.create(
                food=item, qty=qty, total_price=total_price, kitchen=item.kitchen)
            Order.food_orders.add(orderelement)

    Order.save()


def handle_productOrder_element(Order, item, qty):
    total_price = int(qty) * int(item.price)

    try:
        orderelement = Order.product_orders.get(product=item)
        orderelement.qty = int(orderelement.qty) + int(qty)
        orderelement.total_price = int(item.price) * int(orderelement.qty)
        orderelement.isComplete = False

        # If the quantity becomes zero, delete the order element
        if orderelement.qty == 0:
            orderelement.delete()
        else:
            orderelement.save()
    except ObjectDoesNotExist:
        # If the quantity is zero, don't create a new order element
        if qty != 0:
            orderelement = models.ProductOrder.objects.create(
                product=item, qty=qty, total_price=total_price, kitchen=item.kitchen)
            Order.product_orders.add(orderelement)

    Order.save()


class DeliveryOrderAPIView(APIView):

    def post(self, request):
        product_id = request.data.get('pdid')
        qty = request.data.get('qty', 1)
        is_product = request.data.get('ispd', True)

        is_Delivery = request.data.get('isDelivery',True)
        customername = request.data.get('name','')
        address = request.data.get('address','')
        phoneno = request.data.get('phone','')
        except_time = request.data.get('datetime',None)
        description = request.data.get('description','')
        deliveryCharges = request.data.get('deliveryCharges',0)
        description = request.data.get('description','')

        DeliveryOrder, delivery_created = models.DeliveryOrder.objects.get_or_create(
                                            customername=customername,
                                            address=address,
                                            deliveryCharges=deliveryCharges,
                                            exceptTime=except_time,
                                            phoneno=phoneno,
                                            description=description
                                        )

        Order, order_created = models.OrderDetail.objects.get_or_create(
                                            isDelivery=is_Delivery,
                                            deliveryorder=DeliveryOrder
                                        )
        Order.isOrder = False
        Order.save()

        if is_product:
            item = get_object_or_404(models.Product, id=product_id)
            handle_productOrder_element(Order, item, qty)
        else:
            item = get_object_or_404(models.Food, id=product_id)
            handle_foodOrder_element(Order, item, qty)


        try:
            realorder = models.RealOrder.objects.get(orders=Order)

        except ObjectDoesNotExist:
            realorder = models.RealOrder.objects.create(orders=Order)
          
        return Response(status=status.HTTP_200_OK)

class OrderAPIView(APIView):

    def get(self, request):
        table_id = request.GET.get('table_id')
        table = models.Table.objects.get(id=table_id)

        PdOrder = models.OrderDetail.objects.filter(
            table=table, is_paid=False).order_by('-id')

        ser = serializers.OrderDetailsSerializer(PdOrder.first())

        return Response(ser.data)

    def post(self, request):
        print("request.data", request.data)
        table_id = request.data.get('table_id')
        waiter = request.user
        table = get_object_or_404(models.Table, id=table_id)

        product_id = request.data.get('pdid')
        qty = request.data.get('qty', 1)
        is_product = request.data.get('ispd', True)

        PdOrder = models.OrderDetail.objects.filter(
            table=table, is_paid=False).order_by('-id')

        if PdOrder.exists():
            Order = PdOrder.first()
            Order.isOrder = False
            Order.save()    
            table.status = True
            table.save()
        else:
            Order = models.OrderDetail.objects.create(
                table=table, waiter=waiter)
            Order.isOrder = False
            Order.save()
            table.status = True
            table.save()

        if is_product:
            item = get_object_or_404(models.Product, id=product_id)
            handle_productOrder_element(Order, item, qty)
        else:
            item = get_object_or_404(models.Food, id=product_id)
            handle_foodOrder_element(Order, item, qty)

        ser = serializers.OrderDetailsSerializer(Order)
        return Response(ser.data)

    def delete(self, request):
        table_id = request.GET.get('table_id')
        print(table_id)
        table = get_object_or_404(models.Table, id=table_id)

        waiter = request.user

        PdOrder = models.OrderDetail.objects.filter(
            table=table, waiter=waiter, is_paid=False).order_by('-id')

        if PdOrder.exists():
            Order = PdOrder.first()
            Order.product_orders.all().delete()
            Order.food_orders.all().delete()
            Order.save()

        return Response(status=status.HTTP_200_OK)

    def put(self, request):
        itemOrderid = request.data.get('itemorderid')
        isDone = request.data.get('isDone', True)

        print(itemOrderid, isDone)

        order = models.Order.objects.get(id=itemOrderid)
        order.isComplete = isDone
        order.save()

        return Response(status=status.HTTP_200_OK)


class SendOrder(APIView):

    def get(self, request):

        kitchen_id = request.GET.get("kitchen_id")
        time = request.GET.get('time', 'today')
        today = datetime.now()

        this_month_start = today.replace(day=1)
        this_week_start = today - timedelta(days=today.weekday())
        this_year_start = today.replace(month=1, day=1)

        if time == 'today':
            Orders = models.RealOrder.objects.filter(order_time__date=today)
        elif time == 'week':
            Orders = models.RealOrder.objects.filter(
                order_time__date__gte=this_week_start)
        elif time == 'month':
            Orders = models.RealOrder.objects.filter(
                order_time__date__gte=this_month_start)
        elif time == 'year':
            Orders = models.RealOrder.objects.filter(
                order_time__date__gte=this_year_start)

        ser = serializers.RealOrderSerializer(Orders, many=True)

        return Response(ser.data)

    def post(self, request):
        orderdetail_id = request.data.get('order_id')
        guest = request.data.get('guest', 1)
        Order = models.OrderDetail.objects.get(id=orderdetail_id)
        Order.guest = guest
        Order.isOrder = True
        Order.save()
        try:
            realorder = models.RealOrder.objects.get(orders=Order)

        except ObjectDoesNotExist:
            realorder = models.RealOrder.objects.create(orders=Order)

        return Response(status=status.HTTP_200_OK)

    # .................. start cooking........................
    def put(self, request):
        realorder = request.data.get('id')
        put_type = request.data.get('type', 'cook')  # cook, finish
        kitchen = request.data.get('kitchen', 1)

        realorder = models.RealOrder.objects.get(id=realorder)
        if put_type == 'cook':
            realorder.isCooking = True
            for fd in realorder.orders.food_orders.all():
                print(fd.isCooking, "cooking .......")
                fd.isCooking = True
                # if fd.kitchen == kitchen:
                fd.save()

            for pd in realorder.orders.product_orders.all():
                if pd.kitchen == kitchen:
                    if not pd.isCooking:
                        pd.isCooking = True
                    pd.save()

            now = datetime.now()
            realorder.start_cooking_time = now
            realorder.save()

        if put_type == 'finish':
            realorder.isFinish = True
            for fd in realorder.orders.food_orders.all():
                fd.isComplete = True
                fd.save()

            for pd in realorder.orders.product_orders.all():
                pd.isComplete = True
                pd.save()

            now = datetime.now()
            realorder.end_cooking_time = now

            realorder.save()

        return Response(status=status.HTTP_200_OK)

    def delete(self, request):
        # delete also realorder.orders and produt order as sale retun to product
        print(request.data)
        realorder_id = request.GET.get('id')
        realorder = models.RealOrder.objects.get(id=realorder_id)

        realorder.delete()

        return Response(status=status.HTTP_200_OK)

        


def ComputeProductItemFood(food, qty):
    Food = food
    print(food.qty)

    for f in Food.integrient.all():
        pd = f.product
        totalProductUnit = pd.totalunit

        pd.totalunit = int(totalProductUnit) - int((int(f.useunit) * int(qty)))

        pd.qty = int(int(float(pd.totalunit)) / int(float(pd.unit)))

        pd.save()

def ComputeProductItemProduct(product, qty):
    pd = product
    minusunit = int(pd.unit) * int(pd.qty)
    
    pd.totalunit = int(pd.totalunit) - int(minusunit)
    pd.qty = int(int(float(pd.totalunit)) / int(float(pd.unit)))
    pd.save()


class ItemDiscountAPIView(APIView):

    def post(self, request):
        print('item discount', request.data)
        item_ids = request.data.get('item_id', [])
        discount = request.data.get('discount')

        if not item_ids or discount is None:
            return Response({'error': 'Invalid data'}, status=status.HTTP_400_BAD_REQUEST)

        # Assuming the discount is a valid value and the item_ids are valid
        for item_id in item_ids:
            try:
                order = models.Order.objects.get(id=item_id)
                order.discount = discount
                order.save()
            except models.Order.DoesNotExist:
                # Handle the case where an item does not exist, if necessary
                continue

        return Response(status=status.HTTP_201_CREATED)




def MinusOrderRemoveItem(order_item, remove_qty):
    remainqty =  int(remove_qty)
    originalqty  = int(order_item.qty)

    if originalqty > remainqty:
        order_item.qty = originalqty - remainqty
        order_item.save()
        return 0
    elif originalqty == remainqty:
        order_item.delete()
        return 0
    else:
        remainqty = remainqty - originalqty
        order_item.delete()
        return remainqty

def calculateWasteProduct(order_item, ispd, waste_qty):
    if ispd:
        product = order_item.product       
        unit = int(product.unit) * int(waste_qty)
        cost = int(product.cost) * int(waste_qty)      
        models.WasteProduct.objects.create(product=product, unit=unit, cost=cost, description="Waste Product Order ID - #"+ str(order_item.id))
    else:
        food = order_item.food
        integrients = food.integrient.all()
        for i in integrients:
            pd = i.product
            unit = int(i.useunit) * int(waste_qty)
            cost = int(pd.cost) * int(waste_qty)
            models.WasteProduct.objects.create(product=pd, unit=unit, cost=cost, description="Waste Product Order ID - #"+ str(order_item.id))
    


def MinusOrderWasteItem(order_item, waste_qty, ispd=False):
    remainqty =  int(waste_qty)
    originalqty  = int(order_item.qty)

    if originalqty > remainqty:
        order_item.qty = originalqty - remainqty
        calculateWasteProduct(order_item, ispd, waste_qty)
        
        order_item.save()
        return 0
    elif originalqty == remainqty:
        calculateWasteProduct(order_item, ispd, waste_qty)
        order_item.delete()
        return 0
    else:
        remainqty = remainqty - originalqty
        calculateWasteProduct(order_item, ispd, waste_qty)
        order_item.delete()
        return remainqty





def ComputeVoucherData(self, order_ids, orderPayment, isSVH = False, SVH = False):
     AllTotalPrice  = 0
     for order_id in order_ids:
            if isSVH:
                realOrder = models.RealOrder.objects.get(id=order_id)
            else:
                realOrder = models.RealOrder.objects.get(id=order_id.id)

            if isSVH:
                SVH.order.add(realOrder)

            totalPrice = 0
            totalProfit = 0

            for food_order in realOrder.orders.food_orders.all():
                salePriceWithDiscount = int(food_order.food.price) - discountCalculatorWithPerctange(self, food_order.food.price, food_order.discount)
                OriginalProfit = food_order.food.profit

                BuyPrice = int(food_order.food.price) - int(OriginalProfit)
                SaleProfit = int(salePriceWithDiscount) - int(BuyPrice)

                # print(OriginalProfit, "Originalprofit", salePriceWithDiscount, SaleProfit)
                food_order.total_price = int(salePriceWithDiscount) * int(food_order.qty)

                totalPrice += int(salePriceWithDiscount) * int(food_order.qty)
                totalProfit += int(SaleProfit) * int(food_order.qty)
                ComputeProductItemFood(food_order.food,food_order.qty)
                food_order.isPaid = True
                food_order.save()

            for product_order in realOrder.orders.product_orders.all():
                pd = product_order.product
                salePriceWithDiscount = int(pd.price) -  discountCalculatorWithPerctange(self, product_order.product.price, product_order.discount)
                SaleProfit = int(salePriceWithDiscount) - int(pd.cost)
                totalPrice += int(salePriceWithDiscount) * int(product_order.qty)

                product_order.total_price = int(salePriceWithDiscount) * int(product_order.qty)

                totalProfit += int(SaleProfit) * int(product_order.qty)
                ComputeProductItemProduct(pd, product_order.qty)
                product_order.isPaid = True
                product_order.save()
                       
            realOrder.originaltotalPrice = totalPrice
            realOrder.totalProfit = totalProfit
            
            if not realOrder.totalPayment == None:
                realOrder.totalPayment = int(float(realOrder.totalPayment)) + int(float(orderPayment))
           
            realProfit = (convert_to_number(realOrder.totalProfit) + convert_to_number(realOrder.totalPayment)) - int(convert_to_number(realOrder.originaltotalPrice)) - int(convert_to_number(SVH.delivery))
            realOrder.realProfit =  realProfit
            realOrder.isPaid = True
            realOrder.save()
        
            AllTotalPrice += int(totalPrice)
    
     return AllTotalPrice
        

class OrderCompleteAPIView(APIView):
    def get(self, request):
        time = request.GET.get('time', 'today')
        today = datetime.now()

        this_month_start = today.replace(day=1)
        this_week_start = today - timedelta(days=today.weekday())
        this_year_start = today.replace(month=1, day=1)

        if time == 'today':
            Voucher = models.SaveVoucherHistory.objects.filter(
                date__date=today)
        elif time == 'week':
            Voucher = models.SaveVoucherHistory.objects.filter(
                date__date__gte=this_week_start)
        elif time == 'month':
            Voucher = models.SaveVoucherHistory.objects.filter(
                date__date__gte=this_month_start)
        elif time == 'year':
            Voucher = models.SaveVoucherHistory.objects.filter(
                date__date__gte=this_year_start)

        ser = serializers.SaveVoucherHistorySerializer(Voucher, many=True)
        
        return Response(ser.data)
        


    def post(self, request):
        print("request.data",request.data)
        order_ids = request.data.get('order_ids',[])
        table_ids = request.data.get('table_ids',[])

        customername = request.data.get('customername','Unknown')
        payment_type = request.data.get('paymentype','Cash')
        description = request.data.get('description','')
        totalWillPayPrice =  request.data.get('totalPrice',0)
        isDelivery = request.data.get('isDelivery', False)



        discount = request.data.get('discount',0)
        deliveryCharges =  request.data.get('delivery',0)

        guestediscount =  int(float(totalWillPayPrice)) - discountCalculatorWithPerctange(self,totalWillPayPrice,discount) 

        deliveryCharges = convert_to_number(deliveryCharges)
        willtotalPayment = guestediscount + int(float(deliveryCharges))

        totalPayment = request.data.get('totalPayment',willtotalPayment)
        

        if not isDelivery:
            if not table_ids or not order_ids:
                return Response({'error': 'Invalid data'}, status=status.HTTP_400_BAD_REQUEST)

            

        SVH  = models.SaveVoucherHistory.objects.create(customername=customername,
                                    payment_type=payment_type,
                                    description=description,
                                    originaltotalPrice=totalWillPayPrice,
                                    delivery=deliveryCharges,
                                    discount=discount,
                                    totalPayment=totalPayment)

        SVH.totalPrice = int(SVH.originaltotalPrice) - discountCalculatorWithPerctange(self,SVH.originaltotalPrice,discount) + convert_to_number(deliveryCharges)
        

        SVH.save()

        orderPayment = int(SVH.totalPayment) / len(order_ids)
        
        ComputeVoucherData(self, order_ids, orderPayment,True, SVH)
       
        SVH.save()
                
        if not isDelivery:
            for table_id in table_ids:
                table = get_object_or_404(models.Table, id=table_id)
                PdOrder = models.OrderDetail.objects.filter(    
                                table=table).order_by('-id').first()
                PdOrder.is_paid = True
                table.status = False
                PdOrder.save()
                table.save()
        

        return Response(SVH.id)

    def delete(self, request):
        table_id = request.data.get('table_id', None)
        if not table_id == None:
            table = get_object_or_404(models.Table, id=table_id)
            table.status = False
            table.save()
            return Response(status=status.HTTP_200_OK)

        voucher_id = request.data.get('voucher_id',None)
        if not voucher_id == None:
            voucher = models.SaveVoucherHistory.objects.get(id=voucher_id)
            # delete also related realOrders
            realOrders = voucher.order.all()
            for realOrder in realOrders:
                realOrder.delete()

            voucher.delete()


        return Response(status=status.HTTP_200_OK)
    
    def put(self, request):
        voucher_id = request.data.get('voucher_id',None)
        remove_item_ids = request.data.get('remove_item_id',None)
        remove_item_qty =  request.data.get('remove_item_qty',0)
        waste_item_qty = request.data.get('waste_item_qty',0)
        remove_item_type = request.data.get('remove_item_type',"product")
        
        discount = request.data.get('discount',0)
        deliveryCharges =  request.data.get('delivery',0)
        totalPayment = request.data.get('totalPayment',0)

        Voucher =  models.SaveVoucherHistory.objects.get(id=voucher_id)

        realOrders  = Voucher.order.all()
        
        remainqty = remove_item_qty
        remainwasteqty = waste_item_qty
        
        if  not remove_item_ids == None:
            for remove_item_id in remove_item_ids:
        
                for realOrder in realOrders:
                    orderDetail = realOrder.orders

                    if remove_item_type == 'product':
                        product_orders = orderDetail.product_orders.filter(id=remove_item_id)
                        if product_orders.exists():
                            remainqty = MinusOrderRemoveItem(product_orders.first(), remainqty)
                            remainwasteqty = MinusOrderWasteItem(product_orders.first(), waste_item_qty, True)
                    else:
                        food_orders = orderDetail.food_orders.filter(id=remove_item_id)
                        if food_orders.exists():
                            remainqty = MinusOrderRemoveItem(food_orders.first(), remainqty)
                            remainwasteqty = MinusOrderWasteItem(food_orders.first(), waste_item_qty, False)

        if not discount == 0:
            Voucher.discount = discount
        
        if not deliveryCharges == 0:
            Voucher.delivery = deliveryCharges
        
        if not totalPayment == 0:
            Voucher.totalPayment = totalPayment

        Voucher.save()

        orderPayment = int(float(Voucher.totalPayment)) / len(realOrders)


        totalWillPayPrice =  ComputeVoucherData(self, realOrders, orderPayment,False, Voucher)

        guestediscount =  int(float(totalWillPayPrice)) - discountCalculatorWithPerctange(self,totalWillPayPrice,Voucher.discount) 

        deliveryCharges = convert_to_number(Voucher.delivery)
        willtotalPayment = guestediscount + int(float(deliveryCharges))

        Voucher.originaltotalPrice = totalWillPayPrice

        Voucher.totalPrice = willtotalPayment
        Voucher.save()

        return Response(status=status.HTTP_200_OK)        





class ProductPriceChangeWithPercentage(APIView):
    def put(self, request):
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        minus_percentage = request.data.get('minus_perctange', None)
        plus_percentage = request.data.get('plus_perctange', None)

        
        if minus_percentage is not None:
            products = models.Product.objects.filter(user=user)
            for product in products:
                if product.price is not None and minus_percentage is not None:
                    minus_percentage = int(float(minus_percentage))
                    new_price = round(
                        int(float(product.price)) - (int(float(product.price)) * int(float(minus_percentage)) / 100), 2)
                    product.price = int(new_price)
                    product.save()

        if plus_percentage is not None:
            products = models.Product.objects.filter(user=user)
            for product in products:
                if product.price is not None and plus_percentage is not None:
                    new_price = round(
                        int(float(product.price)) + (int(float(product.price)) * int(float(plus_percentage)) / 100), 2)
                    product.price = int(new_price)
                    product.save()

        return Response(status=status.HTTP_201_CREATED)


class ProductPrice(APIView):
    def get(self, request):
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        products = models.Product.objects.filter(user=user)

        # Sum all products cost and multiply with qty to calucate purchase
        total_purchase = 0
        for product in products:
            total_purchase += int(product.cost) * int(product.qty)

        return Response(data=total_purchase, status=status.HTTP_201_CREATED)


def yearGenerator(self, data, strftime='%b'):
    # Create a list of all months in the year

    if strftime == '%b':
        months = [month_abbr[i] for i in range(1, 13)]
    else:
        months = [month_name[i] for i in range(1, 13)]

    # Initialize the result dictionary with all months set to 0
    result = {month: 0 for month in months}

    for month in months:
        print(month)

    for item in data:
        d = datetime.strptime(str(item.date)[0:19], "%Y-%m-%d %H:%M:%S")
        month_name_str = d.strftime(strftime)
        print(month_name_str)
        add_price = float(item.totalProfit)
        result[month_name_str] += add_price

    return result


def monthGenerator(self, data):
    result = {}
    # monthString = ['0','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    # 2022-09-06 18:50:44.169216+00:00

    for item in data:
        d = datetime.strptime(str(item.date)[0:19], "%Y-%m-%d %H:%M:%S")
        name = d.strftime('%x')
        result[name] = result.get(name, 0) + int(float(item.grandtotal))

    return result


def todayGenerator(self, data):
    result = {}
    # monthString = ['0','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    # 2022-09-06 18:50:44.169216+00:00

    for item in data:

        d = datetime.strptime(str(item.date)[0:19], "%Y-%m-%d %H:%M:%S")

        name = d.strftime('%I:%M %p')
        result[name] = result.get(name, 0) + int(float(item.grandtotal))

    return result


def ChartGenerator(self, data, time):
    result = {}
    for item in data:
        data_d = item['date']
        print(data_d, 'Data D')
        d = datetime.strptime(str(data_d), '%Y-%m-%dT%H:%M:%S.%f%z')

        name = d.strftime('%x')
        if time == 'today':
            name = d.strftime('%I:%M %p')
        elif time == 'month':
            name = d.strftime('%x')
        elif time == 'year':
            name = d.strftime('%B')
        else:
            name = d.strftime('%x')
        # print(name)
        result[name] = result.get(name, 0) + int(float(item['totalPayment']))

    return result


def taxCalculatorWithPerctange(self, price, tax_percentage):
    tax_percentage = float(tax_percentage)
    tax = round((float(price) * tax_percentage / 100), 2)
    return tax


def discountCalculatorWithPerctange(self, price, discount_value, isDiscountAmount=False):
    
    if not discount_value :
        discount_value = 0

    if isDiscountAmount:
        return int(float(price) - float(discount_value))

    discount_value = float(discount_value)
    discount = round((float(price) * discount_value / 100), 2)

    print("Discount " , price, discount_value, discount)
    return discount


def checkiinclude(id, soldproducts):
    if not soldproducts == None:
        for i in soldproducts:
            if i['name'] == id:
                return True
            print(i['name'])
        return False
    return True


class Sales(APIView):
    # permission_classes = [AllowAny]P

    def get(self, request):
        type = request.GET.get('type')
        time = request.GET.get('time')
        d = datetime.now()
        print(d)
        chartdata = {}

        if time == 'today':
            data = models.SaveVoucherHistory.objects.filter( date__year=str(
                d.year), date__month=str(d.month), date__day=str(d.day))

        elif time == 'month':
            data = models.SaveVoucherHistory.objects.filter(
                 date__year=str(d.year), date__month=str(d.month))

        elif time == 'year':
            data = models.SaveVoucherHistory.objects.filter(
                 date__year=str(d.year))

        elif time == 'custom':
            start_date = request.GET.get('startd')
            end_date = request.GET.get('endd')
            sd = datetime.strptime(start_date, "%m/%d/%y")
            ed = datetime.strptime(
                end_date, "%m/%d/%y").replace(hour=11, minute=59, second=59)
            data = models.SaveVoucherHistory.objects.filter( date__range=(sd, ed))

        else:
            data = models.SaveVoucherHistory.objects.all()

     
        s = serializers.SaveVoucherHistorySerializer(data, many=True)

        chartdata = ChartGenerator(self, s.data, time)

        # chartdata = todayGenerator(self, s.data)
        # print(s.data,'\n')
        # print(data[0].date)
        # print(chartdata.keys())

        CombineData = {
            'DATA': s.data,
            'CHART_LABEL': chartdata.keys(),
            'CHART_DATA': chartdata.values(),
            'CHART': chartdata,
        }

        return Response(CombineData)

    def post(self, request):
        customerName = request.data['customerName']
        products = request.data['products']
        totalAmount = request.data['totalAmount']
        tax = request.data['tax']
        discount = request.data['discount']
        grandtotal = request.data['grandtotal']
        description = request.data['description']
        deliveryCharges = request.data.get('deliveryCharges', None)
        payment_amount = request.data.get('payment_amount', None)
        isSaveCustomer = request.data.get(
            'isSaveCustomer', 'false').lower() == 'true'
        isDiscountAmount = request.data.get(
            'isDiscountAmount', 'false').lower() == 'true'

        # receiptnumber from sales table and plus one

        print(isDiscountAmount, "Discount amount")

        user = get_user_model().objects.get(username=request.user, is_plan=True)
        last = models.Sales.objects.filter(user=user).last()

        if last is None:
            rn = '1'.zfill(5)
        else:
            rn = str(int(last.voucherNumber) + 1).zfill(5)

        # if isDiscountAmount == 'true':
        #     isDiscountAmount = True

        S = models.Sales.objects.create( customerName=customerName, voucherNumber=rn,
                                        totalAmount=totalAmount, tax=tax, discount=discount, grandtotal=grandtotal, customer_payment=payment_amount,
                                        deliveryCharges=deliveryCharges, isDiscountAmount=isDiscountAmount,
                                        description=description)

        S.save()

        print(products)
        p = json.loads(products)
        print(p)

        totalProfit = 0

        totalAmount = 0  # Each Product

        for b in p:
            print(b)
            # if product is not have in the databse add new product in the database with named catageory to Extra Products

            try:

                product = models.Product.objects.get(id=b['name'], user=user)
                product.qty = int(product.qty) - int(b['qty'])
                product.save()

                profit = (float(b['price']) -
                          float(product.cost)) * float(b['qty'])
                ta = (float(product.price)) * float(b['qty'])

                print("Profit try : ", profit)

                models.SoldProduct.objects.create(
                    name=b['pdname'], price=b['price'], profit=profit, qty=b['qty'],
                    productid=product.id, sales=S, user=user)

            except ObjectDoesNotExist:
                try:
                    category = models.Category.objects.get(
                        title='Extra Products', user=user)
                    if not b.get('cost'):
                        b['cost'] = 0

                    product = models.Product.objects.create(
                        name=b['pdname'], user=user, price=b['price'], cost=int(
                            b['cost']),
                        qty=5, description='Extra Products', category=category)

                    profit = (float(b['price']) -
                              float(product.cost)) * float(b['qty'])
                    ta = (float(product.price)) * float(b['qty'])

                    a = models.SoldProduct.objects.create(
                        name=b['pdname'], price=b['price'], qty=b['qty'],
                        productid=product.id, profit=profit, sales=S, user=user)

                except ObjectDoesNotExist:
                    category = models.Category.objects.create(
                        title='Extra Products', user=user)
                    product = models.Product.objects.create(
                        name=b['pdname'], user=user, price=b['price'], cost=int(
                            b['cost']),
                        qty=5, description='Extra Products', category=category)

                    profit = (float(b['price']) -
                              float(product.cost)) * float(b['qty'])

                    ta = (float(product.price)) * float(b['qty'])

                    a = models.SoldProduct.objects.create(
                        name=b['pdname'], price=b['price'], qty=b['qty'],
                        productid=product.id, profit=profit, sales=S, user=user)

            totalAmount += ta
            totalProfit += profit

        # S.totalProfit = totalProfit - taxCalculatorWithPerctange(self, totalProfit, S.tax)
        S.totalProfit = totalProfit - \
            discountCalculatorWithPerctange(
                self, totalAmount, S.discount, isDiscountAmount)

        S.save()

        saleseri = serializers.SalesSerializer(S)

        if isSaveCustomer:  # == 'true':
            try:
                Customer = models.CustomerName.objects.get(
                    name=customerName, user=user)
                Customer.sales.add(S)
                Customer.save()
            except ObjectDoesNotExist:
                Customer = models.CustomerName.objects.create(
                    name=customerName, user=user)
                Customer.sales.add(S)
                Customer.save()

        return Response(status=status.HTTP_201_CREATED, data=saleseri.data)

    # Delete will be sale return to products and delete from sales

    def delete(self, request):
        id = request.GET.get('id')
        user = get_user_model().objects.get(username=request.user, is_plan=True)

        S = models.Sales.objects.get(user=user, receiptNumber=id)

        Product = models.SoldProduct.objects.filter(sales=S)
        for p in Product:
            try:
                product = models.Product.objects.get(id=p.productid, user=user)
                product.qty = int(product.qty) + int(p.qty)
                product.save()
                p.delete()
            except ObjectDoesNotExist:
                pass
        S.delete()

        return Response(status=status.HTTP_201_CREATED)

    # if user edit the sales voucher like price and qty then it will be updated in the database
    # if user increase product qty then it will be decrease from the Product models qty and if user decrease product qty then it will be increase from the Product models qty
    # then it will be updated in the SoldProduct Database

    def put(self, request):
        print(request.data)
        id = request.data['id']  # receipt Number
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        products = request.data['products']  # old products

        soldproducts = request.data.get('newproducts', None)

        S = models.Sales.objects.get(user=user, receiptNumber=id)
        isSaveCustomer = True

        if float(S.customer_payment) == float(S.grandtotal):
            isSaveCustomer = False

        S.customerName = request.data.get('customerName', S.customerName)
     #   S.totalAmount = totalAmount
        # S.tax = tax
        # S.discount = discount
        # S.grandtotal = grandtotal
        # S.description = description
        # S.deliveryCharges = deliveryCharges

        editedProducts = products

        currentSoldProduct = models.SoldProduct.objects.filter(
            sales=S, user=user)

        ta = 0  # total amount
        pf = 0  # total profit

        # return Response(status=status.HTTP_201_CREATED)

        for currentProduct in currentSoldProduct:
            found = False
            print(currentProduct.id, "Current Product id")
            for editedProduct in editedProducts:

                if currentProduct.id == editedProduct['id']:
                    found = True
                    product = models.Product.objects.get(
                        id=currentProduct.productid, user=user)
                    if int(currentProduct.qty) > int(editedProduct['qty']):
                        product.qty = int(
                            product.qty) + (int(currentProduct.qty) - int(editedProduct['qty']))
                        product.save()
                    elif int(currentProduct.qty) < int(editedProduct['qty']):
                        product.qty = int(
                            product.qty) - (int(editedProduct['qty']) - int(currentProduct.qty))
                        product.save()

                    currentProduct.qty = editedProduct['qty']
                    currentProduct.price = editedProduct['price']
                    pf += (float(currentProduct.price) -
                           float(product.cost)) * float(editedProduct['qty'])
                    ta += (float(currentProduct.price)) * \
                        float(editedProduct['qty'])
                    print(editedProduct['qty'],
                          'Changing edited product : ', ta)

                    currentProduct.profit = (
                        float(currentProduct.price) - float(product.cost)) * float(editedProduct['qty'])
                    currentProduct.save()

                    break

            # check = checkiinclude(currentProduct.productid , soldproducts)
            # if check:
            if not found:
                product = models.Product.objects.get(
                    id=currentProduct.productid, user=user)
                product.qty = int(product.qty) + int(currentProduct.qty)
                product.save()
                currentProduct.delete()

        if not soldproducts == None:
            for newpd in soldproducts:
                pd = models.Product.objects.get(id=newpd['name'], user=user)
                pd.qty = int(pd.qty) - int(newpd['qty'])
                pd.save()

                profit = (float(newpd['price']) -
                          float(pd.cost)) * float(newpd['qty'])

                ta += (float(pd.price)) * float(newpd['qty'])

                a = models.SoldProduct.objects.create(
                    name=pd.name, price=newpd['price'], qty=newpd['qty'],
                    productid=pd.id, profit=profit, sales=S, user=user)

                print(a, "Succes")

                # ta += float(S.totalAmount)
                pf += profit
                S.save()

        S.totalAmount = ta
        S.grandtotal = ta + taxCalculatorWithPerctange(self, ta, S.tax) - discountCalculatorWithPerctange(
            self, ta, S.discount, S.isDiscountAmount) + float(S.deliveryCharges)

        if not isSaveCustomer:
            S.customer_payment = S.grandtotal

        S.totalProfit = pf - \
            discountCalculatorWithPerctange(
                self, ta, S.discount, S.isDiscountAmount)
        S.save()

        return Response(status=status.HTTP_201_CREATED)

def SaveVoucherToRealOrderPayment(S, order_payment):
    print(S.order, order_payment)
    orders = S.order.all()
    if len(orders) > 0:
        order_payment = int(float(order_payment)) / len(orders)

        for realOrder in orders:
            
            if not realOrder.totalPayment == None:
                realOrder.totalPayment = int(float(realOrder.totalPayment)) + int(float(order_payment))

            realProfit = (convert_to_number(realOrder.totalProfit) + convert_to_number(realOrder.totalPayment)) - int(convert_to_number(realOrder.originaltotalPrice))
            realOrder.realProfit =  realProfit
            realOrder.save()
    else:
        print("No orders found.")
  

class CustomerView(APIView):

    def get(self, request):
        cu = models.CustomerName.objects.all()

        ser = serializers.CustomerSerializer(cu, many=True)
        return Response(ser.data)

    def post(self, request):
        customer_name = request.data.get('customerName', None)
        description = request.data.get('description', None)
        cu = models.CustomerName.objects.create(
            name=customer_name, description=description)
        return Response(status=status.HTTP_201_CREATED)

    def put(self, request):
        sale_id = request.data.get('sale_id', None)
        customer_payment = request.data.get('customer_payment', None)


        # add sales to customer
        customer_id = request.data.get("customer_id", None)
        sales_receipt = request.data.get('sales', None)

        # sales data have only saleReceiptNumber [1, 4, 6, 0]

        if not sales_receipt == None:
            C = models.CustomerName.objects.get(id=customer_id)

            for sale in sales_receipt:
                S = models.SaveVoucherHistory.objects.get(id=sale)

                SaveVoucherToRealOrderPayment(S, 0)

      
                S.totalPayment = 0
                S.save()
                C.sales.add(S)
                C.save()

                # Customer Set Payment
        if not sale_id == None:
            if not customer_payment == 0:
                S = models.SaveVoucherHistory.objects.get(id=sale_id)
                SaveVoucherToRealOrderPayment(S, customer_payment)
                S.totalPayment = int(
                    S.totalPayment) + int(customer_payment)
                S.save()

        id = request.data.get('id', None)
        name = request.data.get('name', None)
        description = request.data.get('description', None)

        if not id == None:
            C = models.CustomerName.objects.get(id=id)
            C.name = name
            C.description = description
            C.save()

        return Response(status=status.HTTP_200_OK)

    def delete(self, request):

        customer_id = request.GET.get("customerid", None)
        sales_receipt = request.GET.get('sales', None)


        print(sales_receipt, customer_id)

        if not sales_receipt == None:
            C = models.CustomerName.objects.get(id=customer_id)
            S = models.SaveVoucherHistory.objects.get(
               id=sales_receipt)
            S.save()
            C.sales.remove(S)
            C.save()
        else:
            C = models.CustomerName.objects.get(id=customer_id)
            C.delete()

        return Response(status=status.HTTP_200_OK)


class SupplierView(APIView):

    def get(self, request):
        sup = models.Supplier.objects.all()
        ser = serializers.SupplierSerializer(sup, many=True)
        return Response(ser.data)

    def post(self, request):
        supplier_name = request.data.get('supplierName', None)
        description = request.data.get('description', None)
        cu = models.Supplier.objects.create(
            name=supplier_name, description=description)
        return Response(status=status.HTTP_201_CREATED)

    def put(self, request):
        product_id = request.data.get('product_id', None)
        supplier_payment = request.data.get('supplier_payment', None)


        # add sales to customer
        supplier_id = request.data.get("supplier_id", None)
        productss = request.data.get('products', None)

        # productss data have only saleReceiptNumber [1, 4, 6, 0]

        if not productss == None:
            C = models.Supplier.objects.get(id=supplier_id)

            for pd_id in productss:
                product = models.Product.objects.get(id=pd_id)
                product.supplier_payment = 0
                product.save()
                C.products.add(product)
                C.save()

        # Customer Set Payment
        if not product_id == None:
            if not supplier_payment == 0:
                S = models.Product.objects.get(id=product_id)
                S.supplier_payment = int(
                    S.supplier_payment) + int(supplier_payment)
                S.save()

        id = request.data.get('id', None)
        name = request.data.get('name', None)
        description = request.data.get('description', None)

        if not id == None:
            C = models.Supplier.objects.get(id=id)
            C.name = name
            C.description = description
            C.save()

        return Response(status=status.HTTP_200_OK)

    def delete(self, request):

        supplier_id = request.GET.get("supplier_id", None)
        products = request.GET.get('products', None)


        print(products, supplier_id)

        if not products == None:
            C = models.Supplier.objects.get(id=supplier_id)
            S = models.Product.objects.get(id=products)
            S.save()
            C.products.remove(S)
            C.save()
        else:
            C = models.Supplier.objects.get(id=supplier_id)
            C.delete()

        return Response(status=status.HTTP_200_OK)


class SoldProduct(APIView):

    def get(self, request):
        rn = request.GET.get['receiptNumber']

        user = get_user_model().objects.get(username=request.user, is_plan=True)
        S = models.Sales.objects.get(user=user, receiptNumber=rn)
        seri = serializers.SoldProductSerializer(S.products.all(), many=True)
        return Response(seri.data)


class TopProductsView(APIView):

    def get(self, request):
        time = request.GET.get('time')
        d = datetime.now()

        if time == 'today':
            data = models.OrderDetail.objects.filter( date__year=str(
                d.year), date__month=str(d.month), date__day=str(d.day))

        elif time == 'month':
            data = models.OrderDetail.objects.filter(
                 date__year=str(d.year), date__month=str(d.month))

        elif time == 'year':
            data = models.OrderDetail.objects.filter(
                 date__year=str(d.year))

        elif time == 'custom':
            start_date = request.GET.get('startd')
            end_date = request.GET.get('endd')
            sd = datetime.strptime(start_date, "%m/%d/%y")
            ed = datetime.strptime(
                end_date, "%m/%d/%y").replace(hour=11, minute=59, second=59)
            data = models.OrderDetail.objects.filter(
                 date__range=(sd, ed))

        else:
            data = models.OrderDetail.objects.all()

        topmoneyproduct = {}
        topfreqsellproduct = {}

        for orderDetail in data:          
            for item in orderDetail.food_orders.all():
                topmoneyproduct[item.food.name] = topmoneyproduct.get(
                    item.food.name, 0) + int(float(item.total_price)) * int(float(item.qty))
                topfreqsellproduct[item.food.name] = topfreqsellproduct.get(
                    item.food.name, 0) + 1

        CombineData = {
            'T_Money': topmoneyproduct,
            'T_Freq': topfreqsellproduct,
        }

        return Response(CombineData)


def AyearGenerator(self, data, strftime='%b'):
    # Create a list of all months in the year

    if strftime == '%b':
        months = [month_abbr[i] for i in range(1, 13)]
    else:
        months = [month_name[i] for i in range(1, 13)]

    # Initialize the result dictionary with all months set to 0
    result = {month: 0 for month in months}

    # monthString = ['0','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    # 2022-09-06 18:50:44.169216+00:00

    for item in data:
        if len(str(item.date)) == 10:
            # Date string does not include a time component
            d = datetime.strptime(str(item.date), "%Y-%m-%d")
        else:
            # Date string includes a time component
            d = datetime.strptime(str(item.date)[0:19], "%Y-%m-%d %H:%M:%S")

        month_name_str = d.strftime(strftime)
        result[month_name_str] = result.get(
            month_name_str, 0) + int(float(item.price))

    return result


def AmonthGenerator(self, data):
    result = {}
    # monthString = ['0','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    # 2022-09-06 18:50:44.169216+00:00

    for item in data:
        d = datetime.strptime(str(item.date)[0:19], "%Y-%m-%d")
        name = d.strftime('%x')
        result[name] = result.get(name, 0) + int(float(item.price))

    return result


def AtodayGenerator(self, data):
    result = {}
    # monthString = ['0','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    # 2022-09-06 18:50:44.169216+00:00

    for item in data:
        d = datetime.strptime(str(item.date)[0:19], "%Y-%m-%d")
        name = d.strftime('%I:%M %p')
        result[name] = result.get(name, 0) + int(float(item.price))

    return result


def ChartDataGenerator(self, data, time):
    if time == 'today':
        return AtodayGenerator(self, data)
    elif time == 'month':
        return AmonthGenerator(self, data)
    elif time == 'year':
        return AyearGenerator(self, data)
    else:
        return AmonthGenerator(self, data)


class Expense(APIView):
    def get(self, request):
        time = request.GET.get('time')
        d = datetime.now()
        if time == 'today':
            data = models.Expense.objects.filter( date__year=str(
                d.year), date__month=str(d.month), date__day=str(d.day))
            chartdata = ChartDataGenerator(self, data, time)
        elif time == 'month':
            data = models.Expense.objects.filter(
                 date__year=str(d.year), date__month=str(d.month))
            chartdata = ChartDataGenerator(self, data, time)
        elif time == 'year':
            data = models.Expense.objects.filter(
                 date__year=str(d.year))
            chartdata = ChartDataGenerator(self, data, time)
        elif time == 'custom':
            start_date = request.GET.get('startd')
            end_date = request.GET.get('endd')
            sd = datetime.strptime(start_date, "%m/%d/%y")
            ed = datetime.strptime(
                end_date, "%m/%d/%y").replace(hour=11, minute=59, second=59)
            data = models.Expense.objects.filter(
                date__range=(sd, ed))
            chartdata = ChartDataGenerator(self, data, time)
        else:
            data = models.Expense.objects.all()
            chartdata = ChartDataGenerator(self, data, time)

        s = serializers.ExpenseSerializer(data, many=True)
        CombineData = {
            'DATA': s.data,
            'CHART_LABEL': chartdata.keys(),
            'CHART_DATA': chartdata.values(),
        }

        return Response(CombineData)

    def post(self, request):
        title = request.data['title']
        price = request.data['price']
        date = request.data['date']
        description = request.data['description']

        ex = models.Expense.objects.create(
            date=date,  description=description, title=title, price=price)

        return Response(status=status.HTTP_201_CREATED)

    def put(self, request):
        id = request.data['id']
        title = request.data['title']
        price = request.data['price']
        date = request.data['date']
        description = request.data['description']

        ex = models.Expense.objects.get( id=id)
        ex.title = title
        ex.price = price
        ex.description = description
        ex.date = date
        ex.save()
        return Response(status=status.HTTP_201_CREATED)

    def delete(self, request):
        id = request.GET.get('id')
        ex = models.Expense.objects.get( id=id)
        ex.delete()
        return Response(status=status.HTTP_201_CREATED)


class Purchase(APIView):
    def get(self, request):
        time = request.GET.get('time')
        d = datetime.now()
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        data = models.Purchase.objects.filter(user=user)
        if time == 'today':
            data = models.Purchase.objects.filter( date__year=str(
                d.year), date__month=str(d.month), date__day=str(d.day))
            chartdata = ChartDataGenerator(self, data, time)
        elif time == 'month':
            data = models.Purchase.objects.filter(
                user=user, date__year=str(d.year), date__month=str(d.month))
            chartdata = ChartDataGenerator(self, data, time)
        elif time == 'year':
            data = models.Purchase.objects.filter(
                user=user, date__year=str(d.year))
            chartdata = ChartDataGenerator(self, data, time)
        elif time == 'custom':
            start_date = request.GET.get('startd')
            end_date = request.GET.get('endd')
            sd = datetime.strptime(start_date, "%m/%d/%y")
            ed = datetime.strptime(
                end_date, "%m/%d/%y").replace(hour=11, minute=59, second=59)
            data = models.Purchase.objects.filter(
                user=user, date__range=(sd, ed))
            chartdata = ChartDataGenerator(self, data, time)
        else:
            data = models.Purchase.objects.filter(user=user)
            chartdata = ChartDataGenerator(self, data, time)
        s = serializers.PurchaseSerializer(data, many=True)
        CombineData = {
            'DATA': s.data,
            'CHART_LABEL': chartdata.keys(),
            'CHART_DATA': chartdata.values(),
        }

        return Response(CombineData)

    def post(self, request):
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        title = request.data['title']
        price = request.data['price']
        date = request.data['date']
        description = request.data['description']

        ex = models.Purchase.objects.create(
            date=date, user=user, description=description, title=title, price=price)

        return Response(status=status.HTTP_201_CREATED)

    def put(self, request):
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        id = request.data['id']
        title = request.data['title']
        price = request.data['price']
        date = request.data['date']
        description = request.data['description']

        ex = models.Purchase.objects.get(user=user, id=id)
        ex.title = title
        ex.price = price
        ex.description = description
        ex.date = date
        ex.save()
        return Response(status=status.HTTP_201_CREATED)

    def delete(self, request):
        user = get_user_model().objects.get(username=request.user)
        id = request.GET.get('id')
        ex = models.Purchase.objects.get(user=user, id=id)
        ex.delete()
        return Response(status=status.HTTP_201_CREATED)


class OtherIncome(APIView):
    def get(self, request):
        time = request.GET.get('time')

        d = datetime.now()

        if time == 'today':
            data = models.OtherIncome.objects.filter(date__year=str(
                d.year), date__month=str(d.month), date__day=str(d.day))
            chartdata = ChartDataGenerator(self, data, time)
            print(data)
        elif time == 'month':
            data = models.OtherIncome.objects.filter(
                date__year=str(d.year), date__month=str(d.month))
            chartdata = ChartDataGenerator(self, data, time)
        elif time == 'year':
            data = models.OtherIncome.objects.filter(
                date__year=str(d.year))
            chartdata = ChartDataGenerator(self, data, time)
        elif time == 'custom':
            start_date = request.GET.get('startd')
            end_date = request.GET.get('endd')
            sd = datetime.strptime(start_date, "%m/%d/%y")
            ed = datetime.strptime(
                end_date, "%m/%d/%y").replace(hour=11, minute=59, second=59)
            data = models.OtherIncome.objects.filter(
                date__range=(sd, ed))
            chartdata = ChartDataGenerator(self, data, time)
        else:
            data = models.OtherIncome.objects.all()
            chartdata = ChartDataGenerator(self, data, time)

        s = serializers.OtherIncomeSerializer(data, many=True)
        CombineData = {
            'DATA': s.data,
            'CHART_LABEL': chartdata.keys(),
            'CHART_DATA': chartdata.values(),
        }

        return Response(CombineData)

    def post(self, request):
        title = request.data['title']
        price = request.data['price']
        date = request.data['date']
        description = request.data['description']

        ex = models.OtherIncome.objects.create(
            date=date, description=description, title=title, price=price)

        return Response(status=status.HTTP_201_CREATED)

    def put(self, request):
        id = request.data['id']
        title = request.data['title']
        price = request.data['price']
        date = request.data['date']
        description = request.data['description']

        ex = models.OtherIncome.objects.get(id=id)
        ex.title = title
        ex.price = price
        ex.description = description
        ex.date = date
        ex.save()
        return Response(status=status.HTTP_201_CREATED)

    def delete(self, request):
        id = request.GET.get('id')
        ex = models.OtherIncome.objects.get(id=id)
        ex.delete()
        return Response(status=status.HTTP_201_CREATED)

def OrderyearGenerator(self, data, strftime='%b'):

    if strftime == '%b':
        months = [month_abbr[i] for i in range(1, 13)]
    else:
        months = [month_name[i] for i in range(1, 13)]

    result = {month: 0 for month in months}

    for SVH in data:
        date = SVH.date
        for realorder in SVH.order.all():
            profit = realorder.realProfit
            month_name_str = date.strftime(strftime)
            result[month_name_str] = result.get(
                month_name_str, 0) + int(float(profit))
    print(result)
    return result

class WasteProductView(APIView):

    def get(self, request):
        d = datetime.now()

        time = request.GET.get('time')

        if time == 'today':
            data = models.WasteProduct.objects.filter(
                date__year=str(d.year), date__month=str(d.month), date__day=str(d.day))
        elif time == 'week':
            data = models.WasteProduct.objects.filter(
                date__year=str(d.year), date__week=str(d.isocalendar()[1]))
        elif time == 'month':
            data = models.WasteProduct.objects.filter(
                date__year=str(d.year), date__month=str(d.month))
        elif time == 'year':
            data = models.WasteProduct.objects.filter(
                date__year=str(d.year))
        elif time == 'custom':
            start_date = request.GET.get('startd')
            end_date = request.GET.get('endd')
            sd = datetime.strptime(start_date, "%m/%d/%y")
            ed = datetime.strptime(
                end_date, "%m/%d/%y").replace(hour=11, minute=59, second=59)
            data = models.WasteProduct.objects.filter(
                date__range=(sd, ed))

        s = serializers.WasteProductSerializer(data, many=True)
        
        return Response(s.data)

    def post(self, request):
        product_id = request.data.get('pdid',None)
        unit = request.data.get('unit',0)
        cost = request.data.get('cost',0)
        description = request.data.get('description','')

        pd = models.Product.objects.get(id=product_id)

        wp = models.WasteProduct.objects.create(product=pd, unit=unit, cost=cost, description=description)

        return Response(status=status.HTTP_201_CREATED)

    def delete(self, request):
        id = request.GET.get('id')
        wp = models.WasteProduct.objects.get(id=id)
        wp.delete()
        return Response(status=status.HTTP_201_CREATED)
            


class ProfitAndLoss(APIView):
    def get(self, request, format=None):

        d = datetime.now()

        otherincome_data = models.OtherIncome.objects.filter(
           date__year=str(d.year))
        sales_data = models.SaveVoucherHistory.objects.filter(
           date__year=str(d.year))
        expense_data = models.Expense.objects.filter(
           date__year=str(d.year))
        purchase_data = models.Purchase.objects.filter(
           date__year=str(d.year))

        sales_ge = OrderyearGenerator(self, sales_data, '%B')
        otherincome_ge = AyearGenerator(self, otherincome_data, '%B')
        expense_ge = AyearGenerator(self, expense_data, '%B')
        purchase_ge = AyearGenerator(self, purchase_data, '%B')



        addData = {k: sales_ge.get(k, 0) + otherincome_ge.get(k, 0)
                   for k in set(sales_ge) | set(otherincome_ge)}
        minusData = {k: expense_ge.get(k, 0) + purchase_ge.get(k, 0)
                     for k in set(expense_ge) | set(purchase_ge)}
        subtractData = {k: addData.get(k, 0) - minusData.get(k, 0)
                        for k in set(addData) | set(minusData)}

        print(sales_ge)

        # print(addData,minusData,subtractData)
        ordered_data = sorted(
            addData.items(), key=lambda x: datetime.strptime(x[0], '%B'))
        print(ordered_data)

        CombineData = {
            'addData':  OrderedDict(sorted(addData.items(), key=lambda x: datetime.strptime(x[0], '%B'))),
            'minusData': OrderedDict(sorted(minusData.items(), key=lambda x: datetime.strptime(x[0], '%B'))),
            'result': OrderedDict(sorted(subtractData.items(), key=lambda x: datetime.strptime(x[0], '%B'))),

        }
        return Response(CombineData)


class ProfileAPIView(APIView):

    def get(self, request, format=None):
        user = models.User.objects.get(username=request.user)
        print(user.username)
        s = serializers.ProfileSerializer(user)

        return Response(s.data)

    def post(self, request, format=None):
        user = get_user_model().objects.get(username=request.user)

        if 'image' in request.data:
            user.profileimage = compress_image(request.FILES['image'])
            user.save()
            s = serializers.ProfileSerializer(user)
            return Response(s.data)
        s = serializers.ProfileSerializer(user)
        return Response(s.data)

    def delete(self, request, format=None):
        username = request.GET.get('username')
        user = get_user_model().objects.get(username=request.user)

        if user.is_superuser:
            del_user = get_user_model().objects.get(username=username)
            del_user.delete()

            return Response(status=status.HTTP_200_OK)
        return Response('not access')


class FeedBackAPIView(APIView):

    def post(self, request, format=None):
        message = request.data['message']

        models.FeedBack.objects.create(user=request.user, message=message)

        return Response(status=status.HTTP_201_CREATED)


class PricingAPIView(APIView):
    def get(self, request, format=None):
        data = models.Pricing.objects.filter(is_digits=False)
        pricing_ser = serializers.PricingSerializer(data, many=True)
        user = get_user_model().objects.get(username=request.user)
        print(user.is_superuser)
        pr_req_ser = {data: {}}
        try:
            pricing_req = models.PricingRequest.objects.filter(
                user=user, done=False)
            pr_req_ser = serializers.PricingRequestSerializer(
                pricing_req, many=True)
        except ObjectDoesNotExist:
            print('Objects Does Not exist')
        CombineData = {
            'pricing': pricing_ser.data,
            'pr_request': pr_req_ser.data
        }

        return Response(CombineData)

    # It Is Main Buy SOftware not adding Pricing method check this code
    def post(self, request, format=None):
        price_time_type = request.data['type']
        user = get_user_model().objects.get(username=request.user)
        pricing = models.Pricing.objects.get(
            id=price_time_type, is_digits=False)
        models.PricingRequest.objects.create(user=user, rq_price=pricing)

        return Response(status=status.HTTP_201_CREATED)

    def delete(self, request, format=None):
        price_time_type = request.GET.get('type')
        user = get_user_model().objects.get(username=request.user)
        pricing = models.Pricing.objects.get(
            id=price_time_type, is_digits=False)
        pr_req = models.PricingRequest.objects.get(
            user=user, rq_price=pricing, done=False)
        pr_req.delete()

        return Response(status=status.HTTP_201_CREATED)

# Only Super User Can Be Use this View


class PricingRequestView(APIView):

    def get(self, request, format=None):
        user = get_user_model().objects.get(username=request.user)
        if user.is_superuser:
            pricing_req = models.PricingRequest.objects.all()
            ser_p_r = serializers.PricingRequestSerializer(
                pricing_req, many=True)

            return Response(ser_p_r.data)
        return Response('not access')

    def post(self, request):
        handle_user = get_user_model().objects.get(username=request.user)
        if handle_user.is_superuser:
            username = request.data['username']
            rq_id = request.data['rq_id']
            user = get_user_model().objects.get(username=username)
            pr = models.PricingRequest.objects.get(
                id=rq_id, user=user, done=False)
            pr.done = True
            user.is_plan = True
            start_d = datetime.now()
            end = start_d + timedelta(days=int(pr.rq_price.days))
            print(end)
            user.start_d = start_d  # now Date
            user.end_d = end
            user.save()
            pr.save()
            print(user.start_d)

            return Response(status=status.HTTP_201_CREATED)
        return Response('not access')

    def delete(self, request):
        id = request.GET.get('id')
        user = get_user_model().objects.get(username=request.user)
        if user.is_superuser:
            pricing_req = models.PricingRequest.objects.get(id=id)
            pricing_req.delete()

            return Response(status=status.HTTP_200_OK)
        return Response('not access')


class LogoutUserAPIView(APIView):
    queryset = get_user_model().objects.all()

    def get(self, request, format=None):
        device_unique_id = request.GET.get('duid')
        models.Device.objects.filter(unique_id=device_unique_id).delete()
        request.user.auth_token.delete()
        return Response(status=status.HTTP_200_OK)


class LoginWithFacebook(CreateAPIView):
    permission_classes = [AllowAny]
    serializer_class = serializers.CreateUserSerializer

    def post(self, request):

        try:
            username = request.data['username']
            print(username)
            user = models.User.objects.get(username=username)
            print(user)
            token = Token.objects.get(user=user)
            token_data = {'token': token.key}
            return Response({**token_data}, status=status.HTTP_201_CREATED)
        except ObjectDoesNotExist:
            serializers = self.get_serializer(data=request.data)
            serializers.is_valid(raise_exception=True)
            self.perform_create(serializers)

            token = Token.objects.create(user=serializers.instance)
            token_data = {'token': token.key}

            return Response({**token_data}, status=status.HTTP_201_CREATED)


# Write a class that accept the excel file and read excel file add to product model
# Category row is not id in excel file ,it is title so you have to find the category id and if it not have add new category from row before add product


class ExcelUploadNExportAPIView(APIView):  # for products

    def get(self, request, format=None):
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        type = request.GET.get('type')
        if type == 'all':
            products = models.Product.objects.filter(user=user)
        elif type == 'available':
            products = models.Product.objects.filter(user=user, qty__gt=0)
        elif type == 'out_of_stock':
            products = models.Product.objects.filter(user=user, qty=0)
        # elif type == 'expired':
        #     products = models.Product.objects.filter(user=user,expiry_date__lt=timezone.now())
        else:
            products = models.Product.objects.filter(user=user)

        # Create a new workbook
        wb = Workbook()

        # Add a worksheet for the data
        ws = wb.active
        ws.title = "Products"

        # Define the columns for the worksheet
        columns = ['Product Name', 'Buy Price', 'Sale Price', 'Qty',
                   'Category', 'Description', 'Barcode']

        # Set the column headers for the worksheet
        for col_num, column_title in enumerate(columns, 1):
            ws.cell(row=1, column=col_num,
                    value=column_title).font = Font(bold=True)

        # Add the data to the worksheet
        for row_num, product in enumerate(products, 2):
            ws.cell(row=row_num, column=1, value=product.name)
            ws.cell(row=row_num, column=2, value=product.cost)
            ws.cell(row=row_num, column=3, value=product.price)
            ws.cell(row=row_num, column=4, value=product.qty)
            ws.cell(row=row_num, column=5, value=product.category.title)
            ws.cell(row=row_num, column=6, value=product.description)
            ws.cell(row=row_num, column=7,
                    value=product.barcode)

        # Set the response headers
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=Products.xlsx"

        # Save the workbook to the response
        wb.save(response)

        return response

    def post(self, request, format=None):
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        file = request.FILES['file']
        print(file)
        # Use openpyxl to read the Excel file
        workbook = load_workbook(file, read_only=True)
        sheet = workbook.active

        # Assuming data starts from the second row
        for row in sheet.iter_rows(min_row=2, values_only=True):
            print(row[0])  # Assuming 'Category' is in the first column
            print(row)
            try:
                category = models.Category.objects.get(
                    title=row[4], user=user)
            except ObjectDoesNotExist:
                category = models.Category.objects.create(
                    title=row[4], user=user)

            models.Product.objects.create(
                name=row[0],  # Assuming 'Name' is in the second column
                cost=row[1],  # Assuming "Buy Price " is in the
                price=row[2],  # Assuming 'Price' is in the third column
                qty=row[3],  # Assuming 'Qty' is in the fourth column
                category=category,
                # Assuming 'Description' is in the fifth column
                description=row[5],
                user=user
            )

        return Response(status=status.HTTP_201_CREATED)


class ExcelExportProfitandLoss(APIView):

    def get(self, request, format=None):
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        today = timezone.now()
        # 2022-11-02 08:33:40+00:00
        # endd  = datetime.strptime(str(user.end_d),"%Y-%m-%d %H:%M:%S%z")

        d = datetime.now()

        print(today, 'Today')
        endd = user.end_d
        print(endd, 'End Date')
        print(today >= endd, 'Compare Two Date')
        if today >= endd:
            print('end Plan')
            user.is_plan = False
            user.save()
        else:
            user.is_plan = True
            user.save()

        print(timezone.get_current_timezone)

        s = serializers.ProfileSerializer(user)

        otherincome_data = models.OtherIncome.objects.filter(
            user=user, date__year=str(d.year))
        sales_data = models.Sales.objects.filter(
            user=user, date__year=str(d.year))
        expense_data = models.Expense.objects.filter(
            user=user, date__year=str(d.year))
        purchase_data = models.Purchase.objects.filter(
            user=user, date__year=str(d.year))

        sales_ge = yearGenerator(self, sales_data, '%B')
        otherincome_ge = AyearGenerator(self, otherincome_data, '%B')
        expense_ge = AyearGenerator(self, expense_data, '%B')
        purchase_ge = AyearGenerator(self, purchase_data, '%B')

        addData = {k: sales_ge.get(k, 0) + otherincome_ge.get(k, 0)
                   for k in set(sales_ge) | set(otherincome_ge)}
        minusData = {k: expense_ge.get(k, 0) + purchase_ge.get(k, 0)
                     for k in set(expense_ge) | set(purchase_ge)}
        subtractData = {k: addData.get(k, 0) - minusData.get(k, 0)
                        for k in set(addData) | set(minusData)}

        print(sales_ge)

        # print(addData,minusData,subtractData)
        ordered_data = sorted(
            addData.items(), key=lambda x: datetime.strptime(x[0], '%B'))
        print(ordered_data)

        CombineData = {
            'addData':  OrderedDict(sorted(addData.items(), key=lambda x: datetime
                                           .strptime(x[0], '%B'))),
            'minusData': OrderedDict(sorted(minusData.items(), key=lambda x: datetime
                                            .strptime(x[0], '%B'))),
            'result': OrderedDict(sorted(subtractData.items(), key=lambda x: datetime
                                         .strptime(x[0], '%B'))),
        }
        # Create a new workbook
        wb = Workbook()

        # Add a worksheet for the data
        ws = wb.active
        ws.title = "ProfitandLoss"

        # Define the columns for the worksheet
        columns = ['Month', 'Income', 'Expense', 'Profit/Loss']

        # Set the column headers for the worksheet
        for col_num, column_title in enumerate(columns, 1):
            ws.cell(row=1, column=col_num,
                    value=column_title).font = Font(bold=True)

        # Add the data to the worksheet
        for row_num, month in enumerate(CombineData['addData'], 2):
            ws.cell(row=row_num, column=1, value=month)
            ws.cell(row=row_num, column=2, value=CombineData['addData'][month])
            ws.cell(row=row_num, column=3,
                    value=CombineData['minusData'][month])
            ws.cell(row=row_num, column=4, value=CombineData['result'][month])

        # Set the response headers
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=ProfitandLoss.xlsx"

        # Save the workbook to the response
        wb.save(response)

        return response


# Export All Report includes sales, expenses, purchase, other income to excel file

class ExcelExportAllReportAPIView(APIView):
    def get(self, request, format=None):

        # first get all data report
        type = request.GET.get('type')
        time = request.GET.get('time')
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        d = datetime.now()
        print(d)
        chartdata = {}

        if time == 'today':
            sales_data = models.Sales.objects.filter(user=user, date__year=str(
                d.year), date__month=str(d.month), date__day=str(d.day))
            expense_data = models.Expense.objects.filter(user=user, date__year=str(
                d.year), date__month=str(d.month), date__day=str(d.day))
            purchase_data = models.Purchase.objects.filter(user=user, date__year=str(
                d.year), date__month=str(d.month), date__day=str(d.day))
            otherincome_data = models.OtherIncome.objects.filter(user=user, date__year=str(
                d.year), date__month=str(d.month), date__day=str(d.day))

        elif time == 'month':
            sales_data = models.Sales.objects.filter(
                user=user, date__year=str(d.year), date__month=str(d.month))
            expense_data = models.Expense.objects.filter(
                user=user, date__year=str(d.year), date__month=str(d.month))
            purchase_data = models.Purchase.objects.filter(
                user=user, date__year=str(d.year), date__month=str(d.month))
            otherincome_data = models.OtherIncome.objects.filter(
                user=user, date__year=str(d.year), date__month=str(d.month))
        elif time == 'year':
            sales_data = models.Sales.objects.filter(
                user=user, date__year=str(d.year))
            expense_data = models.Expense.objects.filter(
                user=user, date__year=str(d.year))
            purchase_data = models.Purchase.objects.filter(
                user=user, date__year=str(d.year))
            otherincome_data = models.OtherIncome.objects.filter(
                user=user, date__year=str(d.year))

        elif time == 'custom':
            start_date = request.GET.get('startd')
            end_date = request.GET.get('endd')
            sd = datetime.strptime(start_date, "%m/%d/%y")
            ed = datetime.strptime(
                end_date, "%m/%d/%y").replace(hour=11, minute=59, second=59)
            sales_data = models.Sales.objects.filter(
                user=user, date__range=(sd, ed))
            expense_data = models.Expense.objects.filter(
                user=user, date__range=(sd, ed))
            purchase_data = models.Purchase.objects.filter(
                user=user, date__range=(sd, ed))
            otherincome_data = models.OtherIncome.objects.filter(
                user=user, date__range=(sd, ed))

        else:
            sales_data = models.Sales.objects.filter(user=user)
            expense_data = models.Expense.objects.filter(user=user)
            purchase_data = models.Purchase.objects.filter(user=user)
            otherincome_data = models.OtherIncome.objects.filter(user=user)

        # Make excel file for this all data in different sheet
        # Create a new workbook
        wb = Workbook()

        # Add worksheets for each data type
        sales_ws = wb.create_sheet("Sales")
        expense_ws = wb.create_sheet("Expense")
        otherincome_ws = wb.create_sheet("OtherIncome")

        # Define the columns for each worksheet
        sales_columns = ["Receipt Number", "Customer Name", "Total Amount", "Tax",
                         "Discount", "Delivery Charges", "Grand Total", "Profit", "Date", "Description"]
        expense_columns = ["Title", "Price", "Date", "Description"]
        otherincome_columns = ["Title", "Price", "Date", "Description"]

        # Set the column headers for each worksheet
        for col_num, column_title in enumerate(sales_columns, 1):
            sales_ws.cell(row=1, column=col_num,
                          value=column_title).font = Font(bold=True)
        for col_num, column_title in enumerate(expense_columns, 1):
            expense_ws.cell(row=1, column=col_num,
                            value=column_title).font = Font(bold=True)

        for col_num, column_title in enumerate(otherincome_columns, 1):
            otherincome_ws.cell(row=1, column=col_num,
                                value=column_title).font = Font(bold=True)

        # Add the data to each worksheet
        for row_num, sale in enumerate(sales_data, 2):
            sales_ws.cell(row=row_num, column=1, value=sale.voucherNumber)
            sales_ws.cell(row=row_num, column=2, value=sale.customerName)
            sales_ws.cell(row=row_num, column=3, value=sale.totalAmount)
            sales_ws.cell(row=row_num, column=4, value=sale.tax)
            sales_ws.cell(row=row_num, column=5, value=sale.discount)
            sales_ws.cell(row=row_num, column=6, value=sale.deliveryCharges)
            sales_ws.cell(row=row_num, column=7, value=sale.grandtotal)
            sales_ws.cell(row=row_num, column=8, value=sale.totalProfit)
            sales_ws.cell(row=row_num, column=9,
                          value=sale.date.strftime("%m-%d-%Y"))
            sales_ws.cell(row=row_num, column=10, value=sale.description)

        for row_num, expense in enumerate(expense_data, 2):
            expense_ws.cell(row=row_num, column=1, value=expense.title)
            expense_ws.cell(row=row_num, column=2, value=expense.price)
            expense_ws.cell(row=row_num, column=3,
                            value=expense.date.strftime("%m-%d-%Y"))
            expense_ws.cell(row=row_num, column=4, value=expense.description)

        for row_num, otherincome in enumerate(otherincome_data, 2):
            otherincome_ws.cell(row=row_num, column=1, value=otherincome.title)
            otherincome_ws.cell(row=row_num, column=2, value=otherincome.price)
            otherincome_ws.cell(row=row_num, column=3,
                                value=otherincome.date.strftime("%m-%d-%Y"))
            otherincome_ws.cell(row=row_num, column=4,
                                value=otherincome.description)

        # Set the response headers
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=AllReport.xlsx"

        # Save the workbook to the response
        wb.save(response)

        return response


def generate_barcode(product_id):

    product_id = str(product_id).zfill(12)
    # print(product_id, 'what.....')
    # Generate a unique barcode value using the EAN13 format
    ean = barcode.get_barcode_class('ean13')
    barcode_value = ean(str(product_id), writer=ImageWriter())

    # Convert the barcode image to a base64-encoded string
    buffer = io.BytesIO()
    barcode_value.write(buffer)
    barcode_image = Image.open(buffer)
    return barcode_image

# Export BarCode for all products


class ExcelExportBarCodeAPIView(APIView):
    def get(self, request, format=None):
        user = get_user_model().objects.get(username=request.user, is_plan=True)

        pdlist = request.GET.get('sid')
        pdlist = ast.literal_eval(pdlist)
        product_ids = [int(id) for id in pdlist]
        products = models.Product.objects.filter(user=user, id__in=product_ids)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="Products_BarCode.pdf"'

        # Create a buffer to hold the PDF file
        buffer = io.BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)

        # Set the column and row dimensions
        col_width = 40 * mm
        row_height = 20 * mm
        margin = 10 * mm
        x = margin
        y = A4[1] - margin

        current_product_name = None

        for i in products:
            for j in range(int(i.qty)):
                barcode_image = generate_barcode(i.barcode)

                # Resize the barcode image to fit in the cell
                # barcode_image = barcode_image.resize((int(col_width), int(row_height)))

                # Convert the barcode image to a format that can be added to the PDF
                img_data = io.BytesIO()
                barcode_image.save(img_data, format='PNG')
                img_data.seek(0)

                # c.rect(x, y - row_height - 15 * mm, col_width, row_height, stroke=1, fill=0)

                # Add the barcode image to the PDF
                c.drawImage(ImageReader(img_data), x, y - row_height,
                            width=col_width, height=row_height)

                # Move to the next cell
                x += col_width + 10 * mm

                # If we reach the end of the row, move to the next row
                if x >= A4[0] - margin:
                    x = margin
                    y -= row_height + 10 * mm

                # If we reach the end of the page, start a new page
                if y <= margin:
                    c.showPage()
                    y = A4[1] - margin
                    current_product_name = None

        c.save()

        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)

        return response
